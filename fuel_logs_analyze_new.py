import openpyxl, os, smtplib, sys
from smtplib import SMTP
import pandas as pd
import os.path, time, datetime
from datetime import date
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart

#Varibles
today = datetime.datetime.now().date()
li=[]

#combine the fuel logs from this year (except for today) into 1 excel spreadsheet
for file in os.listdir('G:\\FLEET\\FUEL INFO\\Fuel Export Auto\\'):
    if file.startswith("CompletedExportTemplate_0_2022"): #or file.startswith("test213") #you can change the fileanme here do get more specific dates)
        filename=file
        #print(filename)
        filepath = 'G:\\FLEET\\FUEL INFO\\Fuel Export Auto\\' + filename
        filetime = datetime.datetime.fromtimestamp(
                os.path.getctime('G:\\FLEET\\FUEL INFO\\Fuel Export Auto\\' + file))
        if filetime.date() != today:
            df=pd.read_csv(filepath, usecols =[0,1,3,4,5,], names = ["ID","Gallons","Mileage","Usage","Date",], index_col=False)
            li.append(df)
            frame = pd.concat(li,ignore_index=True)
        else:
            print(file," is from today")

#Export to excel spreadsheet so I can run the same function from the FuelLogClean script            
frame.to_excel('G:\\FLEET\\FUEL INFO\\Fuel Export Auto\\combinedLogs.xlsx', index=False)

#Load XLSX into openpyXL
path = "G:\\FLEET\\FUEL INFO\\Fuel Export Auto\\combinedLogs.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active


#fuel log clean
def clean_data():
    #Take values from column E and put into Column D if necessary
    for i in range(1, 100000):
        # Set variables for different columns  
        ColA = sheet_obj.cell(row=i, column=1)
        ColC = sheet_obj.cell(row=i, column=3)
        ColD = sheet_obj.cell(row=i, column=4)
        # If column 5 has a value then put it in column 4. Exclude certain Equipment ID's (John Deere)
        if ColD.value != 0 and ColA.value != 8406 and ColA.value != 8408 and ColA.value != 8413:
            ColC.value = ColD.value
            
    #delete column 4 (D)
    sheet_obj.delete_cols(4)
    wb_obj.save(path)
   
clean_data()

#sort Columns
def sort_columns():
    #open excel in pandas
    df=pd.read_excel('G:\\FLEET\\FUEL INFO\\Fuel Export Auto\\combinedLogs.xlsx', usecols =[0,1,2,3], names = ["ColA","ColB","ColC","ColD"])
    print(df)
    
    #set variables for columns in spreadsheet
    ColA=df.columns[0]
    ColC=df.columns[2]
    ColD=df.columns[3]
    
    #sort Column A ascending and Column D Descending
    df2=df.sort_values([ColA,ColD],ascending=[True,False])
    
    #save as csv
    df2.to_csv('G:\\FLEET\\FUEL INFO\\Fuel Export Auto\\combinedLogsCleaned.csv', index=False)#, header=None)

    #upload csv into Pandas and add column names
    df3=pd.read_csv('G:\\FLEET\\FUEL INFO\\Fuel Export Auto\\combinedLogsCleaned.csv')

    #Delete 
    df3=df3.dropna(how="all")
    df3.drop(df3[df3['ColD'] == "Date"].index, inplace=True)

    #Set column D to date type and sort
    df3['ColD'] = pd.to_datetime(df3['ColD'])
    df3=df3.sort_values([ColA,ColD],ascending=[True,False])
    df3.to_csv('G:\\FLEET\\FUEL INFO\\Fuel Export Auto\\df3.csv', index=False)

    #Drop duplicate ID's so only the most recent remains
    df_dropped = df3.drop_duplicates(subset=["ColA"], keep = 'first')
    df_dropped2=df_dropped.sort_values([ColA,ColD],ascending=[True,False])
    
    #print(df_dropped2)

    #export back to csv
    df_dropped.to_csv('G:\\FLEET\\FUEL INFO\\Fuel Export Auto\\combinedLogsCleaned.csv', index=False)
    #df_dropped2.to_csv('G:\\FLEET\\FUEL INFO\\Fuel Export Auto\\combinedLogsCleaned2.csv', index=False)
        
sort_columns()

#compare todays log vs combined logs

#todays import
left = pd.read_excel('G:\\FLEET\\FUEL INFO\\Fuel Export Auto\\Fuel Export Master - OMS.xlsx', usecols = [0,1,3,4], names = ["ColA","ColB","ColC","ColD"],header=None, index_col=None)
left['ColA'] = left['ColA'].astype(int)

#print(left)
#left.to_csv('G:\\FLEET\\FUEL INFO\\Fuel Export Auto\\left.csv', index=False)

#combined imports
right = pd.read_csv('G:\\FLEET\\FUEL INFO\\Fuel Export Auto\\combinedLogsCleaned.csv', index_col=None)

#right.drop(right.tail(1).index,inplace=True) # drop last n rows
right['ColA'] = right['ColA'].astype(int)

#print(right)
#right.to_csv('G:\\FLEET\\FUEL INFO\\Fuel Export Auto\\right.csv', index=False)

#join the two imports
result = left.merge(right,how='left',left_on= "ColA", right_on = "ColA",)
#print(result)

#set fields to correct data types for adding and subtracting
result["ColB_x"] = result["ColB_x"].astype(float) #set datatype as float so you can subtract
result["ColB_y"] = result["ColB_y"].astype(float) #set datatype as float so you can subtract
result["ColC_x"] = result["ColC_x"].astype(float) #set datatype as float so you can subtract
result["ColC_y"] = result["ColC_y"].astype(float) #set datatype as float so you can subtract
result["ColD_x"] = pd.to_datetime(result['ColD_x'], infer_datetime_format=True) #set datatype as date so you can subtract
result["ColD_y"] = pd.to_datetime(result['ColD_y'], infer_datetime_format=True) #set datatype as date so you can subtract

#create and calculate the new fields to see difference in fuel, mileage, date of last fuel
result["fuel_diff"] = result["ColB_x"] - result["ColB_y"]
result["mileage_diff"] = result["ColC_x"] - result["ColC_y"]
result["Date_diff"] = result["ColD_x"] - result["ColD_y"]

#Rename Columns
result.rename(columns = {'ColA':'ID','ColB_x':'Todays Fuel', 'ColC_x':'Todays Mileage', 'ColD_x':'Todays Date', 'ColB_y':'Previous Fuel', 'ColC_y':'Previous Mileage','ColD_y':'Previous Date'}, inplace=True)

#Reverse sort so highest mileage difference is on top
result=result.sort_values(result.columns[8], ascending=False)

#Export final to CSV
result.to_csv('G:\\FLEET\\FUEL INFO\\Fuel Export Auto\\comparison.csv', index=False)

#send email with the results in the body
recipients = ['matthew.attiyeh@ci.rosemount.mn.us'] 
emaillist = [elem.strip().split(',') for elem in recipients]
msg = MIMEMultipart()
msg['Subject'] = "Todays Fuel Log Comparison"
msg['From'] = 'GISHelpdesk@ci.rosemount.mn.us'

html = """\
<html>
  <head></head>
  <body>
    {0}
  </body>
</html>
""".format(result.to_html())

part1 = MIMEText(html, 'html')
msg.attach(part1)

server = smtplib.SMTP('cirosemountmnus.mail.protection.outlook.com', 25)
server.sendmail(msg['From'], emaillist , msg.as_string())
